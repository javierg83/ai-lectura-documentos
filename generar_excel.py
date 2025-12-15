import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# DATOS EXACTOS (Sin cabeceras inventadas, transcripción literal)
data = [
    # Fila 1: Se ve cortada arriba. Columna 1 y 2 vacías visualmente, Col 3 tiene texto.
    [
        "", 
        "", 
        "de el o los servicios o suministros finalizados.\n\n(De no ser presentados, los medios de acreditación solicitados, los proyectos informados no serán considerados).\n\nLos proyectos informados que cuenten con un término anticipado de contrato no serán considerados."
    ],
    # Fila 2: Proyecto Privado (Parte superior)
    [
        "", 
        "Proyecto Privado:\nContrato y factura final emitida por los servicios contratados.", 
        "Necesariamente debe adjuntar\nContrato y Factura del último Estado de pago que acredite la ejecución total del suministro informado conjuntamente con “Recepción Conforme o recepción final o certificado emitido por el mandante”.\nSe corroborará su autenticidad a través del Portal Mercado Público."
    ],
    # Fila 3: Proyectos en Ejecución (Esta fila contiene el inicio de la celda fusionada en Columna 1)
    [
        "Proyectos en Ejecución\n\nPara la acreditación de los suministros que se encuentren en ejecución, el oferente deberá acompañar como medios de acreditación y en su totalidad los siguientes antecedentes:", 
        "Proyecto Ámbito público:\nID. Licitación.", 
        "-\tContrato vigente en ejecución mínima de 12 meses,\n-\tLa contratación deberá corresponder a suministro/venta por un monto igual o superior a 100 UTM., según lo establecido en Bases.\n\n(Si estos medios de acreditación no se encuentran publicados en el Portal Mercado Público, los proyectos informados no serán considerados)."
    ],
    # Fila 4: Continuación (La Columna 1 se fusionará con la de arriba)
    [
        "", # Marcador de posición para la fusión
        "Proyecto Privado:\n*contrato vigente y primera y última factura emitida al momento de la presentación de los antecedentes.", 
        "Necesariamente debe adjuntar:\n\n-\tContrato vigente en ejecución mínima de 12 meses,\n-\tLa contratación deberá corresponder a suministro/ventas por un monto igual o superior a 100 UTM., según lo establecido en Bases.\n\n-\tEl oferente deberá presentar la primera factura emitida y ultima factura emitida al momento de la fecha de cierre de este proceso a fin de acreditar el inicio de los servicios y continuidad de estos.\n\n“Estos medios de acreditación, deberán ser proporcionados mediante carpeta digital adjuntos a la oferta de la empresa”.\n(De no ser presentados los proyectos informados no serán considerados)."
    ]
]

# Crear DataFrame sin cabeceras
df = pd.DataFrame(data)
filename = "tabla_exacta_refundida.xlsx"

# Guardar sin header ni index para no ensuciar la tabla
df.to_excel(filename, index=False, header=False)

# Cargar para dar formato
wb = load_workbook(filename)
ws = wb.active

# --- APLICAR REFUNDIDO (MERGE) ---
# Fusionamos la celda A3 con A4 ("Proyectos en Ejecución")
ws.merge_cells('A3:A4')

# --- ESTILOS VISUALES ---
align_style = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Ajuste de anchos para simular la vista del PDF
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 65

# Aplicar bordes y alineación a todo el rango de datos
for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = align_style
        cell.border = thin_border

# Agregar la nota al pie fuera de la tabla principal
ws['A6'] = "NOTA (Se verificará la validez en las facturas en el Servicio de Impuestos Internos)."
ws.merge_cells('A6:C6') # Fusionar nota al pie

wb.save(filename)
print("Excel generado correctamente con celdas refundidas y sin cabeceras inventadas.")